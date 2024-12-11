# Compra el libro en:
# www.ellibrodepython.com

# Instrucciones:
# 1. Instala dependencias: pip install openpyxl
# 2. Ejecuta el script: python 21_manejar_excel.py
# 3. Observa el resultado en el archivo gastos.xlsx

from openpyxl import Workbook
from openpyxl.styles import Font

datos = [
    ["Alquiler", 800],
    ["Electricidad", 100],
    ["Agua", 50],
    ["Internet", 60],
    ["Supermercado", 300]
]

wb = Workbook()
ws = wb.active

ws['A1'] = "Concepto"
ws['B1'] = "Gasto (€)"

bold_font = Font(bold=True)

ws['A1'].font = bold_font
ws['B1'].font = bold_font

for i, (concepto, gasto) in enumerate(datos, start=2):
    ws[f'A{i}'] = concepto
    ws[f'B{i}'] = gasto

ultima = len(datos) + 1

ws[f'A{ultima + 1}'] = "Gastos Totales"
ws[f'B{ultima + 1}'] = f"=SUM(B2:B{ultima})"

wb.save("gastos.xlsx")